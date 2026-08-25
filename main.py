import streamlit as st
import pandas as pd
import unicodedata
import re
from io import BytesIO

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Depurador Observatorios Boyacá",
    layout="wide",
)

ESTILO_CSS = """
<style>
    /* Área principal más compacta */
    .block-container {padding-top: 1.6rem; padding-bottom: 3rem; max-width: 1400px;}

    /* Título principal con degradado institucional */
    .app-title {
        background: linear-gradient(90deg, #14532d 0%, #166534 55%, #15803d 100%);
        color: #ffffff;
        padding: 1.3rem 2rem;
        border-radius: 16px;
        margin-bottom: 0.35rem;
        box-shadow: 0 2px 10px rgba(21, 128, 61, 0.25);
    }
    .app-subtitle {color: #4b5563; margin-bottom: 1.3rem;}

    /* Botones de descarga */
    div[data-testid="stDownloadButton"] button {
        border-radius: 10px;
        font-weight: 600;
        box-shadow: 0 1px 2px rgba(0,0,0,.08);
        transition: transform .08s ease-in-out;
    }
    div[data-testid="stDownloadButton"] button:hover {transform: translateY(-1px);}

    /* Expanders más pulidos */
    details[data-testid="stExpander"] {
        border-radius: 12px;
        border: 1px solid #e5e7eb;
    }

    /* Tarjetas informativas */
    .tarjeta-hero {
        background: linear-gradient(180deg, #f0fdf4 0%, #ffffff 100%);
        border: 1px solid #bbf7d0;
        border-radius: 16px;
        padding: 1.6rem 2rem;
    }
    .paso-tarjeta {
        background: #ffffff;
        border: 1px solid #e5e7eb;
        border-left: 5px solid #16a34a;
        border-radius: 12px;
        padding: 1rem 1.2rem;
        height: 100%;
    }

    /* Barra lateral */
    section[data-testid="stSidebar"] {background: #f8fafc;}
</style>
"""

st.markdown(ESTILO_CSS, unsafe_allow_html=True)




def normalizar_texto(texto) -> str:
    if pd.isna(texto):
        return ""
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        caracter for caracter in texto if not unicodedata.combining(caracter)
    )
    texto = re.sub(r"\s+", " ", texto)
    return texto


def limpiar_columnas_texto(df: pd.DataFrame) -> pd.DataFrame:

    df = df.copy()
    columnas_texto = df.select_dtypes(include=["object", "string"]).columns
    for columna in columnas_texto:
        df[columna] = df[columna].astype("string").str.strip()
        df[columna] = df[columna].replace({"": pd.NA, "nan": pd.NA, "None": pd.NA})
    return df


def limpiar_nombres_columnas(df: pd.DataFrame) -> pd.DataFrame:

    df = df.copy()
    nombres = []
    contador = {}
    for posicion, columna in enumerate(df.columns):
        nombre = str(columna).strip()
        if not nombre or nombre.lower() == "nan":
            nombre = f"columna_{posicion}"
        if nombre in contador:
            contador[nombre] += 1
            nombre_final = f"{nombre}_{contador[nombre]}"
        else:
            contador[nombre] = 0
            nombre_final = nombre
        nombres.append(nombre_final)
    df.columns = nombres
    return df


def convertir_columnas_numericas(df: pd.DataFrame) -> pd.DataFrame:

    df = df.copy()
    for columna in df.columns:
        serie = df[columna]
        if pd.api.types.is_numeric_dtype(serie):
            continue
        if normalizar_texto(columna) in ["DEPARTAMENTO", "DEPTO"]:
            continue

        serie_texto = serie.astype("string").str.strip()

        serie_texto = serie_texto.replace("*", pd.NA)

        cantidad_validos = serie_texto.notna().sum()
        if cantidad_validos == 0:
            continue

        conversion_directa = pd.to_numeric(serie_texto, errors="coerce")
        porcentaje_directo = (
            conversion_directa.notna().sum() / cantidad_validos
            if cantidad_validos > 0
            else 0
        )

        if porcentaje_directo >= 0.70:
            df[columna] = conversion_directa
            continue
        serie_convertida = serie_texto.str.replace(".", "", regex=False)
        serie_convertida = serie_convertida.str.replace(",", ".", regex=False)
        conversion_coma = pd.to_numeric(serie_convertida, errors="coerce")
        porcentaje_coma = (
            conversion_coma.notna().sum() / cantidad_validos
            if cantidad_validos > 0
            else 0
        )

        if porcentaje_coma >= 0.70:
            df[columna] = conversion_coma

    return df


def procesar_csv_deforestacion(contenido: bytes) -> pd.DataFrame:
    codificaciones = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]

    for encoding in codificaciones:
        try:

            df_temp = pd.read_csv(
                BytesIO(contenido),
                encoding=encoding,
                sep=";",
                header=None,
                dtype=str,
                skip_blank_lines=False,
                on_bad_lines="skip",
            )

            df_temp = df_temp.dropna(how="all")

            fila_inicio = None
            for idx, row in df_temp.iterrows():
                row_str = row.astype(str).str.strip()
                if row_str.str.contains("Departamento", case=False, na=False).any():
                    fila_inicio = idx
                    break

            if fila_inicio is None:
                return None

            df_datos = df_temp.iloc[fila_inicio:].copy()

            encabezados = df_datos.iloc[0].astype(str).str.strip().tolist()
            df_datos = df_datos.iloc[1:].copy()

            if len(encabezados) != len(df_datos.columns):

                df_datos.columns = encabezados[: len(df_datos.columns)]
            else:
                df_datos.columns = encabezados

            df_datos = df_datos.dropna(how="all")

            df_datos = limpiar_nombres_columnas(df_datos)
            df_datos = limpiar_columnas_texto(df_datos)

            df_datos = convertir_columnas_numericas(df_datos)

            return df_datos

        except Exception as e:
            continue

    return None


def detectar_fila_cabecera(df_crudo: pd.DataFrame, max_filas: int = 15) -> int:
    mejor_fila = 0
    mejor_puntaje = -1
    limite = min(max_filas, len(df_crudo))
    for fila_indice in range(limite):
        fila = df_crudo.iloc[fila_indice]
        no_vacias = fila.notna() & (fila.astype(str).str.strip() != "")
        textos = fila[no_vacias].astype(str).str.strip()
        if textos.empty:
            continue
        unicidad = textos.nunique() / len(textos)
        puntaje = no_vacias.sum() * unicidad
        if puntaje > mejor_puntaje:
            mejor_fila = fila_indice
            mejor_puntaje = puntaje
    return mejor_fila


@st.cache_data(show_spinner=False)
def construir_dataframe(df_crudo: pd.DataFrame, fila_cabecera: int) -> pd.DataFrame:
    if fila_cabecera >= len(df_crudo):
        return pd.DataFrame()
    encabezados = df_crudo.iloc[fila_cabecera].astype(str).str.strip()
    df = df_crudo.iloc[fila_cabecera + 1 :].copy()
    df.columns = encabezados
    df = limpiar_nombres_columnas(df)
    df = df.reset_index(drop=True)
    df = limpiar_columnas_texto(df)
    df = convertir_columnas_numericas(df)
    return df




@st.cache_data(show_spinner=False)
def leer_excel(contenido: bytes, hoja: str) -> pd.DataFrame:
    return pd.read_excel(BytesIO(contenido), sheet_name=hoja, header=None)


@st.cache_data(show_spinner=False)
def leer_csv_estandar(contenido: bytes) -> pd.DataFrame:

    codificaciones = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]
    separadores = [",", ";", "\t", "|"]

    for encoding in codificaciones:
        for sep in separadores:
            try:
                return pd.read_csv(
                    BytesIO(contenido),
                    encoding=encoding,
                    sep=sep,
                    engine="python",
                    header=None,
                    dtype=str,
                    skip_blank_lines=False,
                )
            except (UnicodeDecodeError, pd.errors.ParserError):
                continue

    raise ValueError("No fue posible interpretar el CSV.")


@st.cache_data(show_spinner=False)
def obtener_hojas_excel(contenido: bytes):
    excel = pd.ExcelFile(BytesIO(contenido))
    return excel.sheet_names





@st.cache_data(show_spinner=False)
def detectar_columnas_departamento(df: pd.DataFrame):

    candidatas = []
    for columna in df.columns:
        columna_normalizada = normalizar_texto(columna)
        if "DEPARTAMENTO" in columna_normalizada or "DEPTO" in columna_normalizada:
            candidatas.append(columna)


    if not candidatas:
        for columna in df.columns:
            if df[columna].dtype == "object" or df[columna].dtype == "string":
                muestra = df[columna].dropna().astype(str).head(20)
                if any("BOYACA" in normalizar_texto(val) for val in muestra):
                    candidatas.append(columna)

    return candidatas


@st.cache_data(show_spinner=False)
def filtrar_boyaca(df: pd.DataFrame, columna_departamento: str) -> pd.DataFrame:

    rango_combinantes = (
        "[" + "".join(chr(codigo) for codigo in range(0x0300, 0x0370)) + "]"
    )
    normalizado = (
        df[columna_departamento]
        .astype("string")
        .str.strip()
        .str.upper()
        .str.normalize("NFKD")
        .str.replace(rango_combinantes, "", regex=True)
        .str.replace(r"\s+", " ", regex=True)
        .fillna("")
    )
    return df.loc[normalizado == "BOYACA"].copy()





@st.cache_data(show_spinner=False)
def detectar_columnas_periodo(df: pd.DataFrame) -> list:

    columnas_periodo = []

    for columna in df.columns:
        texto = normalizar_texto(columna)


        if re.fullmatch(r"\d{4}", texto):
            año = int(texto)
            if 1900 <= año <= 2100:
                columnas_periodo.append(columna)

        
        elif re.match(r"\d{4}-\d{4}", texto):
            columnas_periodo.append(columna)

        
        elif re.match(r"\d{4}\s*-\s*\d{4}", texto):
            columnas_periodo.append(columna)


    def extraer_anio_inicio(col):
        texto = normalizar_texto(col)
        if re.match(r"\d{4}", texto):
            return int(re.search(r"\d{4}", texto).group())
        return 0

    columnas_periodo.sort(key=extraer_anio_inicio)

    return columnas_periodo





@st.cache_data(show_spinner=False)
def preparar_datos_comparacion(
    df: pd.DataFrame, eje_x: str, ejes_y: list
) -> pd.DataFrame:

    columnas = [eje_x] + ejes_y
    datos = df[columnas].copy()

    for col in ejes_y:
        if pd.api.types.is_numeric_dtype(datos[col]):
            datos[col] = datos[col].interpolate(limit_direction="both")
    return datos





@st.cache_data(show_spinner=False)
def preparar_datos_evolucion(
    df: pd.DataFrame,
    columnas_periodo: list,
    columna_categoria: str,
    valores_categoria: list,
) -> pd.DataFrame:

    if not valores_categoria:
        return pd.DataFrame()

    datos_base = df[
        df[columna_categoria]
        .astype(str)
        .isin([str(valor) for valor in valores_categoria])
    ].copy()

    if datos_base.empty:
        return pd.DataFrame()

    # Asegurarse de que las columnas de periodo existen
    columnas_existentes = [col for col in columnas_periodo if col in datos_base.columns]
    if not columnas_existentes:
        return pd.DataFrame()

    datos = datos_base.melt(
        id_vars=[columna_categoria],
        value_vars=columnas_existentes,
        var_name="Periodo",
        value_name="Valor",
    )

    
    def extraer_anio(texto):
        if pd.isna(texto):
            return None
        texto = str(texto)
        match = re.search(r"\d{4}", texto)
        if match:
            return int(match.group())
        return None

    datos["Año_inicio"] = datos["Periodo"].apply(extraer_anio)
    datos = datos.dropna(subset=["Año_inicio", "Valor"])

    
    datos["Valor"] = pd.to_numeric(datos["Valor"], errors="coerce")
    datos = datos.dropna(subset=["Valor"])

    
    datos = datos.sort_values(by=["Año_inicio", columna_categoria])
    datos = datos.reset_index(drop=True)

    
    datos["Periodo"] = datos["Periodo"].astype(str)

    return datos





def preparar_datos_grafica(
    df: pd.DataFrame,
    modo: str,
    eje_x: str = None,
    ejes_y: list = None,
    columnas_periodo: list = None,
    columna_categoria: str = None,
    valores_categoria: list = None,
) -> tuple:

    if modo == "Comparar categorías":
        datos = preparar_datos_comparacion(df, eje_x, ejes_y)
        return (datos, eje_x, ejes_y)

    if modo == "Evolución temporal":
        datos = preparar_datos_evolucion(
            df=df,
            columnas_periodo=columnas_periodo,
            columna_categoria=columna_categoria,
            valores_categoria=valores_categoria,
        )
        return (datos, "Periodo", ["Valor"])

    return (pd.DataFrame(), None, [])




def construir_grafica(
    datos: pd.DataFrame,
    tipo: str,
    eje_x: str,
    ejes_y: list,
    modo: str = "Comparar categorías",
    columna_categoria: str = None,
):

    if datos.empty:
        st.warning("No hay datos para graficar")
        return

    if modo == "Comparar categorías":
        if tipo == "Línea":
            st.line_chart(datos, x=eje_x, y=ejes_y, width="stretch")
        elif tipo == "Barra":
            st.bar_chart(datos, x=eje_x, y=ejes_y, width="stretch")
        elif tipo == "Área":
            st.area_chart(datos, x=eje_x, y=ejes_y, width="stretch")
        elif tipo == "Dispersión":
            st.scatter_chart(datos, x=eje_x, y=ejes_y, width="stretch")
        return

    if modo == "Evolución temporal":
        
        if columna_categoria is None or columna_categoria not in datos.columns:
            
            datos_grafico = datos.set_index("Periodo")[["Valor"]]
            if tipo == "Línea":
                st.line_chart(datos_grafico, width="stretch")
            elif tipo == "Barra":
                st.bar_chart(datos_grafico, width="stretch")
            elif tipo == "Área":
                st.area_chart(datos_grafico, width="stretch")
            elif tipo == "Dispersión":
                st.scatter_chart(datos_grafico, width="stretch")
            return

        
        datos_pivot = datos.pivot_table(
            index="Periodo", columns=columna_categoria, values="Valor", aggfunc="first"
        )

        
        def extraer_anio(texto):
            if pd.isna(texto):
                return 0
            match = re.search(r"\d{4}", str(texto))
            return int(match.group()) if match else 0

        
        datos_pivot["_orden"] = datos_pivot.index.map(extraer_anio)
        datos_pivot = datos_pivot.sort_values("_orden")
        datos_pivot = datos_pivot.drop(columns=["_orden"])

        columnas_series = [col for col in datos_pivot.columns if col != "Periodo"]
        if not columnas_series:
            st.warning("No existen series suficientes para construir la gráfica.")
            return

        
        datos_grafico = datos_pivot.reset_index()

        if tipo == "Línea":
            st.line_chart(
                datos_grafico, x="Periodo", y=columnas_series, width="stretch"
            )
        elif tipo == "Barra":
            st.bar_chart(datos_grafico, x="Periodo", y=columnas_series, width="stretch")
        elif tipo == "Área":
            st.area_chart(
                datos_grafico, x="Periodo", y=columnas_series, width="stretch"
            )
        elif tipo == "Dispersión":
            st.scatter_chart(
                datos_grafico, x="Periodo", y=columnas_series, width="stretch"
            )





def limpiar_nombre_archivo(nombre: str) -> str:
    nombre = normalizar_texto(nombre)
    nombre = re.sub(r"[^A-Z0-9_.-]", "_", nombre)
    return nombre




MIME_EXCEL = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def nombre_hoja_valido(nombre: str) -> str:
    
    seguro = re.sub(r"[\[\]:*?/\\]", "_", str(nombre)).strip()
    return (seguro or "Datos")[:31]


@st.cache_data(show_spinner=False, max_entries=32)
def generar_csv_bytes(df: pd.DataFrame) -> bytes:
    
    return df.to_csv(index=False).encode("utf-8-sig")


def aplicar_formato_tabla(hoja, df: pd.DataFrame, max_filas_estilo: int = 3000) -> None:
    
    if df.empty or len(df.columns) == 0:
        return

    relleno = PatternFill("solid", fgColor="166534")
    fuente_blanca = Font(color="FFFFFF", bold=True, size=11)
    linea = Side(style="thin", color="D1D5DB")
    borde = Border(left=linea, right=linea, top=linea, bottom=linea)

    total_columnas = len(df.columns)


    limite_presupuesto = (
        max(1, 60_000 // total_columnas) if total_columnas else max_filas_estilo
    )
    total_filas = min(len(df) + 1, max_filas_estilo, limite_presupuesto)


    for indice in range(1, total_columnas + 1):
        celda = hoja.cell(row=1, column=indice)
        celda.fill = relleno
        celda.font = fuente_blanca
        celda.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        celda.border = borde
    hoja.row_dimensions[1].height = 26


    for fila in range(2, total_filas + 1):
        for indice in range(1, total_columnas + 1):
            hoja.cell(row=fila, column=indice).border = borde

 
    for indice, columna in enumerate(df.columns, start=1):
        muestra = df[columna].head(300).dropna().astype(str)
        ancho_contenido = max((len(valor) for valor in muestra), default=0)
        ancho = min(max(len(str(columna)) + 2, ancho_contenido + 2), 45)
        hoja.column_dimensions[get_column_letter(indice)].width = ancho


    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = f"A1:{get_column_letter(total_columnas)}{total_filas}"


@st.cache_data(show_spinner=False, max_entries=16)
def generar_excel_bytes(df: pd.DataFrame, nombre_hoja: str = "Datos") -> bytes:
    
    hoja = nombre_hoja_valido(nombre_hoja)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as escritor:
        df.to_excel(escritor, index=False, sheet_name=hoja)
        aplicar_formato_tabla(escritor.sheets[hoja], df)
    return buffer.getvalue()


@st.cache_data(show_spinner=False, max_entries=8)
def generar_excel_libro(hojas: tuple) -> bytes:
    
    buffer = BytesIO()
    usados = {}
    with pd.ExcelWriter(buffer, engine="openpyxl") as escritor:
        for nombre_fuente, df in hojas:
            hoja = nombre_hoja_valido(nombre_fuente)
            base, contador = hoja, 1
            while hoja in usados:
                contador += 1
                sufijo = f"_{contador}"
                hoja = base[: 31 - len(sufijo)] + sufijo
            usados[hoja] = df
            df.to_excel(escritor, index=False, sheet_name=hoja)
        for hoja, df in usados.items():
            aplicar_formato_tabla(escritor.sheets[hoja], df)
    return buffer.getvalue()


def mostrar_panel_descargas(
    df: pd.DataFrame,
    nombre_base: str,
    key: str,
    titulo: str = "Descargar esta tabla",
) -> None:
    
    with st.container(border=True):
        col_titulo, col_csv, col_excel = st.columns([3, 1.1, 1.1])
        with col_titulo:
            st.markdown(f"##### {titulo}")
            st.caption(
                f"{len(df)} filas × {len(df.columns)} columnas · el archivo conserva "
                "las columnas y valores exactamente como aparecen en la tabla."
            )
        with col_csv:
            st.download_button(
                "CSV",
                icon=":material/download:",
                data=generar_csv_bytes(df),
                file_name=f"{limpiar_nombre_archivo(nombre_base)}.csv",
                mime="text/csv",
                key=f"{key}_csv",
                help="Compatible con Excel y otras herramientas.",
                use_container_width=True,
            )
        with col_excel:
            st.download_button(
                "Excel",
                icon=":material/download:",
                data=generar_excel_bytes(df, nombre_base),
                file_name=f"{limpiar_nombre_archivo(nombre_base)}.xlsx",
                mime=MIME_EXCEL,
                key=f"{key}_excel",
                type="primary",
                help="Con formato: encabezados, anchos ajustados y autofiltros.",
                use_container_width=True,
            )



with st.sidebar:
    st.markdown("### Panel de control")
    st.caption("Sube una matriz Excel o CSV para comenzar.")
    archivo = st.file_uploader(
        "Selecciona un archivo Excel o CSV",
        type=["xlsx", "xls", "csv"],
        help="Formatos aceptados: .xlsx, .xls y .csv",
    )
    if archivo:
        st.success(f"**{archivo.name}**", icon=":material/description:")

    st.divider()
    with st.container(border=True):
        st.markdown("**Flujo de trabajo**")
        st.markdown(
            "1. Carga el archivo  \n"
            "2. Selecciona las hojas  \n"
            "3. Filtra Boyacá y descarga  \n"
            "4. Grafica y exporta"
        )


st.markdown(
    '<div class="app-title">'
    '<h1 style="margin:0;color:#ffffff;display:flex;align-items:center;gap:.65rem;">'
    '<svg width="28" height="28" viewBox="0 0 24 24" fill="#ffffff" '
    'xmlns="http://www.w3.org/2000/svg">'
    '<rect x="4" y="12" width="4" height="9"/>'
    '<rect x="10" y="7" width="4" height="14"/>'
    '<rect x="16" y="3" width="4" height="18"/></svg>'
    "Depurador Observatorios Boyacá</h1></div>",
    unsafe_allow_html=True,
)
st.markdown(
    '<p class="app-subtitle">Procesamiento, filtrado y exportación de matrices Excel y CSV</p>',
    unsafe_allow_html=True,
)


if archivo is None:
    with st.container(border=True):
        st.markdown('<div class="tarjeta-hero">', unsafe_allow_html=True)
        st.markdown("### Bienvenido")
        st.markdown(
            "Esta herramienta depura matrices de datos de los observatorios de Boyacá: "
            "detecta la cabecera automáticamente, filtra los registros del departamento, "
            "permite **descargarlos con formato** y construir gráficas listas para analizar."
        )
        st.markdown("</div>", unsafe_allow_html=True)

        p1, p2, p3 = st.columns(3)
        with p1:
            st.markdown(
                '<div class="paso-tarjeta"><h4 style="margin-top:0;">Paso 1 · Cargar</h4>'
                "Sube tu archivo Excel o CSV desde el <b>panel de la izquierda</b>.</div>",
                unsafe_allow_html=True,
            )
        with p2:
            st.markdown(
                '<div class="paso-tarjeta"><h4 style="margin-top:0;">Paso 2 · Depurar</h4>'
                "Selecciona hojas, confirma la cabecera y filtra los registros de <b>Boyacá</b>.</div>",
                unsafe_allow_html=True,
            )
        with p3:
            st.markdown(
                '<div class="paso-tarjeta"><h4 style="margin-top:0;">Paso 3 · Descargar</h4>'
                "Exporta cualquier tabla en <b>CSV o Excel</b>, igual que se ve en pantalla.</div>",
                unsafe_allow_html=True,
            )
    st.stop()

if archivo:
    contenido = archivo.getvalue()
    extension = archivo.name.lower().rsplit(".", 1)[-1]
    if extension in ("xlsx", "xls"):
        tipo_archivo = "excel"
        try:
            hojas_disponibles = obtener_hojas_excel(contenido)
        except Exception as error:
            st.error(f"No fue posible leer el archivo Excel: {error}")
            st.stop()

        st.header("2. Seleccionar hojas")
        with st.container(border=True):
            hojas_seleccionadas = st.multiselect(
                "Selecciona una o varias hojas",
                options=hojas_disponibles,
                help="Puedes seleccionar dos o más hojas del archivo.",
            )

            if not hojas_seleccionadas:
                st.info("Selecciona al menos una hoja.")
                st.stop()

    elif extension == "csv":
        tipo_archivo = "csv"
        st.header("2. Archivo CSV")

        with st.container(border=True):

            es_deforestacion = st.checkbox(
                "¿Es el archivo de deforestación del IDEAM? (con ; como separador)",
                value=True,
            )

            if es_deforestacion:
                st.info("Usando procesamiento especial para archivos de deforestación.")
                hojas_seleccionadas = ["CSV_Deforestacion"]
            else:
                hojas_seleccionadas = ["CSV"]
                st.success("Archivo CSV cargado correctamente.")
    else:
        st.error("Formato de archivo no compatible.")
        st.stop()




    resultados = {}
    datos_procesados = False

    for nombre_hoja in hojas_seleccionadas:
        st.divider()
        st.subheader(f"Fuente: {nombre_hoja}")


        try:
            if tipo_archivo == "excel":
                df_crudo = leer_excel(contenido, nombre_hoja)


                fila_sugerida = detectar_fila_cabecera(df_crudo)

                with st.expander("Configuración de cabecera"):
                    st.caption(
                        "Revisa las primeras filas y confirma cuál contiene los nombres de las columnas."
                    )
                    st.dataframe(df_crudo.head(15), width="stretch")

                    fila_cabecera = st.number_input(
                        "Fila de cabecera (0 = primera fila)",
                        min_value=0,
                        max_value=min(14, len(df_crudo) - 1),
                        value=fila_sugerida,
                        step=1,
                        key=f"cabecera_{nombre_hoja}",
                    )

                df = construir_dataframe(df_crudo, fila_cabecera)
                if df.empty:
                    st.error("No quedaron datos después de construir el DataFrame.")
                    continue

            elif tipo_archivo == "csv" and "Deforestacion" in nombre_hoja:

                df = procesar_csv_deforestacion(contenido)
                if df is None:
                    st.error("No se pudo procesar el archivo de deforestación.")
                    continue

                st.success("Archivo de deforestación procesado correctamente.")

                with st.expander("Vista previa de los datos"):
                    st.dataframe(df.head(20), width="stretch")
                    st.caption(
                        f"Total de filas: {len(df)} | Total de columnas: {len(df.columns)}"
                    )

            else:

                df_crudo = leer_csv_estandar(contenido)
                fila_sugerida = detectar_fila_cabecera(df_crudo)

                with st.expander("Configuración de cabecera"):
                    st.caption(
                        "Revisa las primeras filas y confirma cuál contiene los nombres de las columnas."
                    )
                    st.dataframe(df_crudo.head(15), width="stretch")

                    fila_cabecera = st.number_input(
                        "Fila de cabecera (0 = primera fila)",
                        min_value=0,
                        max_value=min(14, len(df_crudo) - 1),
                        value=fila_sugerida,
                        step=1,
                        key=f"cabecera_{nombre_hoja}",
                    )

                df = construir_dataframe(df_crudo, fila_cabecera)
                if df.empty:
                    st.error("No quedaron datos después de construir el DataFrame.")
                    continue

            candidatas_depto = detectar_columnas_departamento(df)

            if candidatas_depto:
                col_depto = st.selectbox(
                    "Columna Departamento", candidatas_depto, key=f"depto_{nombre_hoja}"
                )
            else:
                st.warning("No se encontró automáticamente la columna Departamento.")
                col_depto = st.selectbox(
                    "Selecciona manualmente la columna Departamento",
                    df.columns,
                    key=f"depto_manual_{nombre_hoja}",
                )



            df_boyaca = filtrar_boyaca(df, col_depto)

            if df_boyaca.empty:
                st.error(f"No se encontraron registros de Boyacá en '{nombre_hoja}'.")


                with st.expander("Ver valores encontrados en la columna Departamento"):
                    valores_unicos = (
                        df[col_depto].dropna().astype(str).unique().tolist()
                    )
                    st.write(f"Total de valores únicos: {len(valores_unicos)}")
                    st.write(sorted(valores_unicos))
                continue

            st.success(f"Se encontraron {len(df_boyaca)} registros de Boyacá.")


            periodos_boyaca = detectar_columnas_periodo(df_boyaca)
            m1, m2, m3 = st.columns(3)
            m1.metric("Registros Boyacá", f"{len(df_boyaca):,}".replace(",", "."))
            m2.metric("Columnas", len(df_boyaca.columns))
            m3.metric("Periodos detectados", len(periodos_boyaca))


            with st.expander("Ver datos de Boyacá", expanded=len(df_boyaca) <= 25):
                st.dataframe(df_boyaca, width="stretch")



            mostrar_panel_descargas(
                df=df_boyaca,
                nombre_base=f"Boyaca_{nombre_hoja}",
                key=f"desc_boyaca_{nombre_hoja}",
                titulo="Descargar registros filtrados de Boyacá",
            )

            st.markdown("### Filtrar datos")


            columnas_texto = df_boyaca.select_dtypes(
                include=["object", "string"]
            ).columns.tolist()

            columnas_texto = [col for col in columnas_texto if col != col_depto]

            if not columnas_texto:
                st.info(
                    "No hay columnas de texto adicionales para filtrar. Usando todos los datos."
                )
                df_filtrado = df_boyaca.copy()
            else:
                col_filtro = st.selectbox(
                    "Columna para filtrar (ej: Municipio)",
                    columnas_texto,
                    key=f"filtro_{nombre_hoja}",
                )

                opciones = sorted(
                    df_boyaca[col_filtro].dropna().astype(str).unique().tolist()
                )
                valores = st.multiselect(
                    "Valores que deseas analizar",
                    opciones,
                    default=opciones[:1] if opciones else [],
                    key=f"valores_{nombre_hoja}",
                )

                if not valores:
                    st.info("Selecciona al menos un valor para filtrar.")
                    df_filtrado = df_boyaca.copy()
                else:
                    df_filtrado = df_boyaca[
                        df_boyaca[col_filtro].astype(str).isin(valores)
                    ].copy()

            quitar_duplicados = st.checkbox(
                "Quitar filas duplicadas", value=True, key=f"duplicados_{nombre_hoja}"
            )
            duplicados_eliminados = 0
            if quitar_duplicados:
                filas_antes = len(df_filtrado)
                df_filtrado = df_filtrado.drop_duplicates()
                duplicados_eliminados = filas_antes - len(df_filtrado)

            if df_filtrado.empty:
                st.warning("No quedaron datos después de aplicar los filtros.")
                continue

            st.success(f"Datos finales: {len(df_filtrado)} filas.")

            r1, r2, r3 = st.columns(3)
            r1.metric("Registros finales", f"{len(df_filtrado):,}".replace(",", "."))
            r2.metric("Registros Boyacá previos", len(df_boyaca))
            r3.metric("Duplicados eliminados", duplicados_eliminados)

            mostrar_panel_descargas(
                df=df_filtrado,
                nombre_base=f"Boyaca_{nombre_hoja}_filtrado",
                key=f"desc_filtrado_{nombre_hoja}",
                titulo="Descargar datos con todos los filtros aplicados",
            )

            st.markdown("### Configuración de gráfica")


            columnas_periodo = detectar_columnas_periodo(df_filtrado)

            if columnas_periodo:
                st.success(f"Periodos detectados: {', '.join(columnas_periodo)}")
            else:
                st.warning("No se detectaron periodos (años o rangos de años).")

            modos_disponibles = ["Comparar categorías"]
            if columnas_periodo:
                modos_disponibles.append("Evolución temporal")

            modo_analisis = st.segmented_control(
                "Modo de análisis",
                modos_disponibles,
                default=modos_disponibles[0],
                key=f"modo_{nombre_hoja}",
            )
            if modo_analisis is None:
                modo_analisis = modos_disponibles[0]

            col_x = None
            cols_y = []
            columna_categoria = None
            valores_categoria = []

            if modo_analisis == "Comparar categorías":
                st.caption(
                    "Utiliza una columna como Eje X y una o varias columnas numéricas como Eje Y."
                )

                sugerencias_x = [col_depto] + [
                    col for col in df_filtrado.columns if col not in columnas_periodo
                ]
                col_x = st.selectbox(
                    "Eje X (categoría)", sugerencias_x, key=f"x_{nombre_hoja}"
                )

                columnas_numericas = df_filtrado.select_dtypes(
                    include="number"
                ).columns.tolist()

                columnas_numericas = [
                    col
                    for col in columnas_numericas
                    if df_filtrado[col].notna().sum() > 0
                ]

                if not columnas_numericas:
                    st.warning(
                        "No existen columnas numéricas para utilizar como Eje Y."
                    )
                    continue

                cols_y = st.multiselect(
                    "Eje Y (valores numéricos)",
                    columnas_numericas,
                    default=(
                        columnas_numericas[:2]
                        if len(columnas_numericas) >= 2
                        else columnas_numericas
                    ),
                    key=f"y_{nombre_hoja}",
                )

                if not cols_y:
                    st.info("Selecciona al menos una columna para el Eje Y.")
                    continue

            else:
                st.caption("Visualiza la evolución de los valores a través del tiempo.")


                columnas_identificacion = [
                    col for col in df_filtrado.columns if col not in columnas_periodo
                ]

                if not columnas_identificacion:
                    st.warning("No existen columnas de identificación para las series.")
                    continue

                columna_categoria = st.selectbox(
                    "Columna que identifica la serie (ej: Municipio)",
                    columnas_identificacion,
                    key=f"categoria_{nombre_hoja}",
                )

                opciones_categoria = sorted(
                    df_filtrado[columna_categoria]
                    .dropna()
                    .astype(str)
                    .unique()
                    .tolist()
                )

                valores_categoria = st.multiselect(
                    "Series que deseas visualizar",
                    opciones_categoria,
                    default=(
                        opciones_categoria[:3]
                        if len(opciones_categoria) >= 3
                        else opciones_categoria
                    ),
                    key=f"series_{nombre_hoja}",
                )

                if not valores_categoria:
                    st.info("Selecciona al menos una serie.")
                    continue

                col_x = "Periodo"
                cols_y = ["Valor"]

            tipo_grafica = st.segmented_control(
                "Tipo de gráfica",
                ["Línea", "Barra", "Área", "Dispersión"],
                default="Línea",
                key=f"grafica_{nombre_hoja}",
            )
            if tipo_grafica is None:
                tipo_grafica = "Línea"

            datos_grafica, eje_x_final, ejes_y_final = preparar_datos_grafica(
                df=df_filtrado,
                modo=modo_analisis,
                eje_x=col_x,
                ejes_y=cols_y,
                columnas_periodo=columnas_periodo,
                columna_categoria=columna_categoria,
                valores_categoria=valores_categoria,
            )

            if datos_grafica.empty:
                st.warning("No existen datos suficientes para construir la gráfica.")
                continue

            with st.expander("Ver datos utilizados para la gráfica"):
                st.dataframe(datos_grafica, width="stretch")
                st.caption(f"Total de registros: {len(datos_grafica)}")

            if modo_analisis == "Evolución temporal":
                periodos = datos_grafica["Periodo"].unique()
                st.caption(
                    f"Periodos: {min(periodos) if len(periodos) > 0 else 'N/A'} → "
                    f"{max(periodos) if len(periodos) > 0 else 'N/A'} | "
                    f"Series: {', '.join(valores_categoria[:5])}{' ...' if len(valores_categoria) > 5 else ''}"
                )
            else:
                st.caption(
                    f"Eje X: {eje_x_final} | "
                    f"Ejes Y: {', '.join(ejes_y_final[:3])}{' ...' if len(ejes_y_final) > 3 else ''}"
                )

            st.markdown(f"### Gráfica: {tipo_grafica}")
            construir_grafica(
                datos=datos_grafica,
                tipo=tipo_grafica,
                eje_x=eje_x_final,
                ejes_y=ejes_y_final,
                modo=modo_analisis,
                columna_categoria=(
                    columna_categoria if modo_analisis == "Evolución temporal" else None
                ),
            )

            resultado = datos_grafica.copy()
            resultado.insert(0, "HOJA", nombre_hoja)
            if modo_analisis == "Evolución temporal" and columna_categoria:
                resultado.insert(1, "CATEGORIA", columna_categoria)
            resultados[nombre_hoja] = resultado
            datos_procesados = True

        except Exception as error:
            st.error(f"No fue posible procesar '{nombre_hoja}': {error}")
            import traceback

            st.code(traceback.format_exc())
            continue

if resultados:
    st.divider()
    st.header("3. Exportar resultados")

    df_final = pd.concat(resultados.values(), ignore_index=True)

    e1, e2, e3 = st.columns(3)
    e1.metric("Fuentes procesadas", len(resultados))
    e2.metric("Registros totales", f"{len(df_final):,}".replace(",", "."))
    e3.metric("Columnas combinadas", len(df_final.columns))

    with st.expander("Vista previa consolidada", expanded=True):
        st.dataframe(df_final.head(300), width="stretch", height=300)
        total_formateado = f"{len(df_final):,}".replace(",", ".")
        if len(df_final) > 300:
            st.caption(
                f"Mostrando 300 de {total_formateado} registros · "
                "la descarga incluye todos."
            )
        else:
            st.caption(f"Total de registros exportados: {len(df_final)}")

    col_excel, col_csv = st.columns(2)
    with col_excel:
        st.download_button(
            "Descargar todo en Excel (.xlsx)",
            icon=":material/grid_on:",
            data=generar_excel_libro(tuple(resultados.items())),
            file_name=limpiar_nombre_archivo(
                f"Boyaca_resultados_{len(resultados)}_fuentes.xlsx"
            ),
            mime=MIME_EXCEL,
            type="primary",
            use_container_width=True,
            help="Un libro con una hoja formateada por cada fuente procesada.",
        )
    with col_csv:
        st.download_button(
            "Descargar todo en CSV",
            icon=":material/table_chart:",
            data=generar_csv_bytes(df_final),
            file_name=limpiar_nombre_archivo(
                f"Boyaca_resultados_{len(resultados)}_fuentes.csv"
            ),
            mime="text/csv",
            use_container_width=True,
            help="Todos los registros consolidados en un solo archivo UTF-8.",
        )

    with st.expander(f"Descargar cada fuente por separado ({len(resultados)})"):
        for nombre_fuente, df_fuente in resultados.items():
            st.markdown(f"**{nombre_fuente}**")
            f1, f2 = st.columns(2)
            with f1:
                st.download_button(
                    "CSV",
                    icon=":material/table_chart:",
                    data=generar_csv_bytes(df_fuente),
                    file_name=limpiar_nombre_archivo(f"Boyaca_{nombre_fuente}.csv"),
                    mime="text/csv",
                    key=f"export_{nombre_fuente}_csv",
                    use_container_width=True,
                )
            with f2:
                st.download_button(
                    "Excel",
                    icon=":material/grid_on:",
                    data=generar_excel_bytes(df_fuente, nombre_fuente),
                    file_name=limpiar_nombre_archivo(f"Boyaca_{nombre_fuente}.xlsx"),
                    mime=MIME_EXCEL,
                    key=f"export_{nombre_fuente}_excel",
                    use_container_width=True,
                )
            st.divider()
